"""Excel → CSV data-merge prep + InDesign Data Merge driver.

Workflow:
1. The user has multiple .xlsx files where each WORKSHEET is one US state.
   Row 1 = headers, row 2 = that state's data values, row 3+ = optional
   secondary tables (per-state programs/partners etc.) we don't merge.
2. We flatten across all files: one row per state, columns aligned.
3. Column names are sanitized to InDesign-safe placeholder names
   (no spaces, no punctuation, lowercase_snake_case).
4. Output:
     - data_merge.csv      → fed to InDesign's Data Merge panel
     - placeholders.md     → human-readable mapping the user uses
                             when tagging the template
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import List


_SHORT_NAME_RULES = [
    # (full-name regex, short token) — applied in order
    (r"^state(\b|_area)?$",                                   "state"),
    (r"^total_corporate_investments_and_giving$",             "total"),
    (r"^investment_capital_investment_in.*facilities$",       "capital"),
    (r"^investment_capital",                                  "capital"),
    (r"^capital_in$",                                         "capital"),
    (r"^investment_contract_manufacturing.*",                 "cmo"),
    (r"^contract_manufacturing_organizations$",               "cmo"),
    (r"^investment_contract_research.*",                      "cro"),
    (r"^contract_research_organizations_and_inve.*",          "cro"),
    (r"^giving_corporate_grants_and_awards_progr.*",          "giving"),
    (r"^giving_corporate_grants_and_awards.*",                "giving"),
    (r"^clinical_trials_programs_clinical_trials.*",          "ct_total"),
    (r"^clinical_trials_programs_oncology.*",                 "ct_onco"),
    (r"^clinical_trials_programs_inflammation$",              "ct_inflam"),
    (r"^clinical_trials_programs_virology$",                  "ct_virol"),
    (r"^oncology_and_cell_therapy$",                          "ct_onco"),
    (r"^facilities$",                                         "facilities"),
    (r"^workforce_workers$",                                  "wf_total"),
    (r"^workforce_full_time.*",                               "wf_ft"),
    (r"^workforce_contingent.*",                              "wf_ct"),
    (r"^top_20_investments.*original.*",                      "top20_orig"),
    (r"^top_20_investments.*",                                "top20"),
    (r"^cro_and_research$",                                   "top20"),
    (r"^corporate_grants_and_awards_programs$",               "grant_prog"),
    (r"^corporate_grants_and_awards_programs_value$",         "grant_val"),
    (r"^value$",                                              "grant_val"),
    (r"^area_specific_grants_hiv.*",                          "g_hiv"),
    (r"^area_specific_grants_liver.*",                        "g_liv"),
    (r"^area_specific_grants_oncology.*",                     "g_onc"),
    (r"^area_specific_grants_other.*",                        "g_oth"),
    (r"^racial_equity_community.*",                           "racial_eq"),
    (r"^compass.*",                                           "compass"),
]


def _sanitize_column_name(raw: str) -> str:
    """Convert a column header into an InDesign-safe placeholder name.

    Two-stage:
      1. Long-form: drop punctuation, snake_case, lowercase. Used as a
         deterministic key for the rule table below.
      2. Short-form: map the long-form to a compact token (≤ 12 chars)
         so InDesign data-merge placeholders don't blow up the layout
         (a 44-char placeholder overflows a frame designed for a 12-char
         dollar value, breaking the layout during editing).

    Returns the SHORT name — that's what becomes the InDesign field name
    via the CSV header row.
    """
    s = (raw or "").strip()
    s = re.sub(r"\s+", " ", s)
    # Detect distinguishing parenthetical qualifiers BEFORE we strip parens.
    # The Excel has two near-identical Top 20 columns: one tagged "(Original)"
    # with UPPERCASE values, another (without that qualifier) with the
    # cleaned-up title-case values we actually want.
    if re.search(r"top\s*20.+\(\s*original\s*\)", s, re.I):
        return "top20_orig"
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[-:,&]", " ", s)
    long_form = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_").lower()
    if not long_form:
        return ""
    # Stage 2: rule-based shortening
    for pattern, short in _SHORT_NAME_RULES:
        if re.match(pattern, long_form):
            return short
    # Generic fallback — first ~12 chars of the long form
    if len(long_form) <= 14:
        return long_form
    # Drop common filler then truncate
    cleaned = long_form
    for filler in ("clinical_trials_programs_", "area_specific_grants_",
                   "corporate_grants_and_awards_programs_", "workforce_",
                   "investment_", "top_20_investments_"):
        cleaned = cleaned.replace(filler, "")
    return cleaned[:14].strip("_")


def _read_sheet_data(ws):
    """Return (headers, data_row) from a single worksheet. Both lists of strings."""
    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    if len(rows) < 2:
        return [], []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data    = [str(c) if c is not None else "" for c in rows[1]]
    return headers, data


# Columns whose values are stored as a vertical list across multiple rows
# in the worksheet (one entry per row). The flattener gathers all non-empty
# rows for these columns and joins them with newlines so the merge step can
# drop the whole list into a single placeholder.
_LIST_COLUMNS = {"top20", "grant_prog"}


def _read_sheet_lists(ws, headers):
    """Walk every row of ws and gather non-empty values for any column whose
    sanitized header matches a name in _LIST_COLUMNS. Returns
    {sanitized_col: "line1\\nline2\\n..."}.
    """
    # Map each list-column sanitized-name to the column index in the sheet
    col_idx_for = {}
    for idx, h in enumerate(headers):
        sn = _sanitize_column_name(h)
        if sn in _LIST_COLUMNS and sn not in col_idx_for:
            col_idx_for[sn] = idx
    if not col_idx_for:
        return {}
    buckets = {sn: [] for sn in col_idx_for}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        for sn, ci in col_idx_for.items():
            if ci < len(row):
                v = row[ci]
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    buckets[sn].append(s)
    return {sn: "\n".join(vals) for sn, vals in buckets.items() if vals}


def _build_unified_schema(xlsx_paths: List[str]):
    """Walk all xlsx sheets and pick a single canonical column ordering.
    Returns a list of (sanitized_name, original_header) — order is the order
    they first appear in the first non-empty sheet.
    """
    import openpyxl
    canonical = []           # list of (sanitized, original)
    seen_sanitized = set()
    for xp in xlsx_paths:
        wb = openpyxl.load_workbook(xp, data_only=True, read_only=True)
        for ws in wb.worksheets:
            headers, _ = _read_sheet_data(ws)
            for col_idx, h in enumerate(headers):
                if not h.strip():
                    continue
                sn = _sanitize_column_name(h)
                if not sn or sn in seen_sanitized:
                    continue
                # Deduplicate name collisions
                base = sn
                n = 2
                while sn in seen_sanitized:
                    sn = f"{base}_{n}"
                    n += 1
                seen_sanitized.add(sn)
                canonical.append((sn, h.strip()))
        wb.close()
    return canonical


def _build_maps_index(maps_folder: str | Path) -> dict:
    """Walk a folder of pre-made map files and return a dict
    keyed by NORMALIZED state name (lowercase, no spaces/punctuation)
    → absolute path. Convention: filename starts with the state name,
    e.g. 'California SEP Map.ai', 'NewMexico SEP Map.ai'. Recursive.
    """
    if not maps_folder:
        return {}
    folder = Path(maps_folder)
    if not folder.exists() or not folder.is_dir():
        return {}
    index = {}
    for p in folder.rglob("*"):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        if ext not in (".ai", ".pdf", ".eps", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".psd"):
            continue
        stem = p.stem
        # Take the first chunk before "SEP", "Map", "_", etc. as the state guess
        # then normalize. Just being generous — we'll do prefix matches.
        norm = re.sub(r"[^A-Za-z]+", "", stem).lower()
        if norm and norm not in index:
            index[norm] = str(p.resolve())
    return index


def _match_state_to_map(state_name: str, maps_index: dict) -> str:
    """Find a map path for a state. Tolerates 'New Mexico'/'NewMexico',
    'Washington D.C.'/'DC', 'Puerto Rico'/'PR'. Returns empty string if no
    match — InDesign Data Merge tolerates blank image paths (frame stays empty).
    """
    if not maps_index or not state_name:
        return ""
    norm_state = re.sub(r"[^A-Za-z]+", "", state_name).lower()
    if not norm_state:
        return ""
    # Direct prefix match — most filenames begin with the state name
    for key, path in maps_index.items():
        if key.startswith(norm_state):
            return path
    # Fallback: state name appears anywhere in the filename
    for key, path in maps_index.items():
        if norm_state in key:
            return path
    # Common abbreviations
    aliases = {
        "districtofcolumbia": "dc",
        "puertorico": "pr",
    }
    alt = aliases.get(norm_state)
    if alt:
        for key, path in maps_index.items():
            if key.startswith(alt):
                return path
    return ""


def flatten_xlsx_to_csv(xlsx_paths: List[str], out_csv: str | Path,
                       out_placeholders_md: str | Path = None,
                       maps_folder: str | Path = None) -> dict:
    """Read every state worksheet across `xlsx_paths` and produce a single
    CSV with one row per state. Returns a summary dict.

    If `maps_folder` is provided, an extra `@image` column is added to the CSV
    where each value is the absolute path to that state's map file (matched
    by normalized state name). The `@` prefix is the InDesign Data Merge
    convention for image fields — the user drags `@image` from the Data Merge
    panel onto an image frame in the template, and at merge time InDesign
    places the per-record image automatically.
    """
    import openpyxl

    out_csv = Path(out_csv)
    canonical = _build_unified_schema(xlsx_paths)
    headers_by_sanitized = {sn: orig for sn, orig in canonical}
    column_order = [sn for sn, _ in canonical]
    maps_index = _build_maps_index(maps_folder) if maps_folder else {}

    state_rows = {}  # name → dict of sanitized→value
    state_order = []
    for xp in xlsx_paths:
        wb = openpyxl.load_workbook(xp, data_only=True, read_only=True)
        for ws in wb.worksheets:
            headers, data = _read_sheet_data(ws)
            if not headers or not data:
                continue
            state_name = (data[0] if data else "").strip() or ws.title.strip()
            if state_name in state_rows:
                continue  # first occurrence wins
            row = {}
            for col_idx, h in enumerate(headers):
                if not h.strip() or col_idx >= len(data):
                    continue
                sn = _sanitize_column_name(h)
                if not sn or sn not in column_order:
                    continue
                row[sn] = data[col_idx].strip()
            # Override list-typed columns with the multi-row joined values
            list_vals = _read_sheet_lists(ws, headers)
            for sn, joined in list_vals.items():
                row[sn] = joined
            # Always set the state name under a known key
            row.setdefault("state", state_name)
            state_rows[state_name] = row
            state_order.append(state_name)
        wb.close()

    # Ensure 'state' is the leading column
    if "state" not in column_order:
        column_order = ["state"] + column_order
        headers_by_sanitized["state"] = "State / Area"

    # If a maps folder was provided, add a `@image` image column AND populate
    # each row with the matched path. Track which states had no match for
    # the summary.
    map_matched = []
    map_missing = []
    if maps_index:
        column_order = column_order + ["@image"]
        headers_by_sanitized["@image"] = "Map (image field)"
        for sn in state_order:
            mp = _match_state_to_map(sn, maps_index)
            state_rows[sn]["@image"] = mp
            (map_matched if mp else map_missing).append(sn)

    # Format known numeric columns: $-prefix for monetary, comma-separated
    # for big counts. Done at write-time so the merged InDesign file shows
    # values like "$8,787,744" instead of raw "8787744".
    _MONEY_COLS = {"total", "capital", "cmo", "cro", "giving", "grant_val",
                   "g_hiv", "g_liv", "g_onc", "g_oth"}
    _COUNT_COLS = {"wf_total", "wf_ft", "wf_ct"}
    def _fmt_value(col: str, raw: str) -> str:
        s = (raw or "").strip()
        if not s or s.upper() == "NA":
            return s
        if col in _MONEY_COLS or col in _COUNT_COLS:
            try:
                # Accept "8787744", "8787744.0", "8,787,744" etc.
                cleaned = s.replace(",", "").replace("$", "")
                num = float(cleaned)
                if num == 0:
                    return "0"
                if num == int(num):
                    formatted = f"{int(num):,}"
                else:
                    formatted = f"{num:,.2f}"
                return ("$" + formatted) if col in _MONEY_COLS else formatted
            except (ValueError, TypeError):
                return s
        return s

    for sn in state_order:
        row = state_rows[sn]
        for col in list(row.keys()):
            if col == "@image" or col == "state":
                continue
            row[col] = _fmt_value(col, row.get(col, ""))

    # Write CSV
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=column_order, extrasaction="ignore")
        w.writeheader()
        for sn in state_order:
            w.writerow(state_rows[sn])

    # Write placeholders.md (the user's tagging cheat sheet)
    if out_placeholders_md:
        out_md = Path(out_placeholders_md)
        out_md.parent.mkdir(parents=True, exist_ok=True)
        with open(out_md, "w", encoding="utf-8") as f:
            f.write("# Data Merge — Placeholder Reference\n\n")
            f.write(f"Generated from: {', '.join(Path(p).name for p in xlsx_paths)}\n")
            f.write(f"Total states: {len(state_order)}\n\n")
            f.write("## How to use\n\n")
            f.write("1. Open the InDesign template.\n")
            f.write("2. **Window → Utilities → Data Merge** to open the panel.\n")
            f.write("3. From the panel menu, choose **Select Data Source…** and pick the generated `data_merge.csv`.\n")
            f.write("4. Drag each placeholder below from the Data Merge panel onto the matching number/text in the template.\n")
            if maps_index:
                f.write("5. **For the map**: drag the `@image` placeholder onto the image frame. ")
                f.write("InDesign will auto-place the matching state's map at merge time.\n")
                f.write("6. Save the template.\n")
                f.write("7. Re-run the app's data merge — it'll generate one .indd per state.\n\n")
            else:
                f.write("5. Save the template.\n")
                f.write("6. Re-run the app's data merge — it'll generate one .indd per state.\n\n")
            f.write("## Placeholder ↔ Source column\n\n")
            f.write("| Placeholder (drag from panel) | Source column |\n")
            f.write("|---|---|\n")
            for sn in column_order:
                orig = headers_by_sanitized.get(sn, "?")
                f.write(f"| `<<{sn}>>` | {orig} |\n")
            f.write("\n## Sample values (California row)\n\n")
            sample = state_rows.get("California") or next(iter(state_rows.values()), {})
            f.write("| Placeholder | Value |\n|---|---|\n")
            for sn in column_order:
                v = sample.get(sn, "")
                f.write(f"| `<<{sn}>>` | {v} |\n")

    return {
        "csv_path": str(out_csv),
        "placeholders_md": str(out_placeholders_md) if out_placeholders_md else None,
        "states": state_order,
        "columns": column_order,
        "n_columns": len(column_order),
        "n_states": len(state_order),
        "n_maps_matched": len(map_matched) if maps_index else 0,
        "maps_missing": map_missing,
    }
